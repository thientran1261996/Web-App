import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def find_max(self):
    max = self[0]
    for i in self:
        if max < i:
            max = i
    print(max)


def find_min(self):
    min = self[0]
    for i in self:
        if min > i:
            min = i
    print(min)


def make_list():
    count = int(input("Input your list length: "))
    i = 0
    list = []
    while i < count:
        list[i] = input()
        i += 1
    return list


def process_excel(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 4).value = round(sheet.cell(row, 3).value * 0.9, 2)
    sheet.cell(1, 4).value = 'new'
    values = Reference(sheet,
        max_col=4,
        max_row=sheet.max_row,
        min_col=4,
        min_row=2
    )
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e1')
    wb.save("transaction)2.xlsx")